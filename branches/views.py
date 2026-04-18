import csv
import io
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_GET, require_http_methods

from openpyxl import Workbook, load_workbook

from .forms import BranchForm, BranchImportForm, StaffSignupForm
from .models import Branch

staff_required = user_passes_test(
    lambda u: u.is_authenticated and u.is_staff,
    login_url="login",
)


def home(request):
    return render(request, "branches/home.html")


@require_http_methods(["GET", "POST"])
def staff_signup(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect("dashboard")

    if request.method == "POST":
        form = StaffSignupForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = True
            user.save()
            login(request, user)
            messages.success(request, "Staff account created and signed in.")
            return redirect("dashboard")
        messages.error(request, "Please correct the errors below.")
    else:
        form = StaffSignupForm()

    return render(request, "registration/signup.html", {"form": form})


@require_GET
def branch_search_api(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 1:
        return JsonResponse({"results": []})
    qs = (
        Branch.objects.filter(
            models.Q(branch_name__icontains=q) | models.Q(contact_number__icontains=q)
        )[:15]
        .values("id", "branch_name", "contact_number")
    )
    return JsonResponse({"results": list(qs)})


@require_GET
def branch_detail_api(request):
    pk = request.GET.get("id")
    if not pk:
        return JsonResponse({"error": "id required"}, status=400)
    branch = get_object_or_404(Branch, pk=pk)
    return JsonResponse(
        {
            "id": branch.id,
            "branch_name": branch.branch_name,
            "contact_number": branch.contact_number,
            "address": branch.address or "",
        }
    )


@staff_required
def dashboard(request):
    branches = Branch.objects.all()
    import_form = BranchImportForm()
    return render(
        request,
        "branches/dashboard.html",
        {"branches": branches, "import_form": import_form},
    )


@staff_required
@require_http_methods(["GET", "POST"])
def branch_add(request):
    if request.method == "POST":
        form = BranchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch added successfully.")
            return redirect("dashboard")
        messages.error(request, "Please correct the errors below.")
    else:
        form = BranchForm()
    return render(request, "branches/add_branch.html", {"form": form})


@staff_required
@require_http_methods(["GET", "POST"])
def branch_edit(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == "POST":
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch updated successfully.")
            return redirect("dashboard")
        messages.error(request, "Please correct the errors below.")
    else:
        form = BranchForm(instance=branch)
    return render(
        request,
        "branches/edit_branch.html",
        {"form": form, "branch": branch},
    )


@staff_required
@require_http_methods(["GET", "POST"])
def branch_delete(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == "POST":
        name = branch.branch_name
        branch.delete()
        messages.success(request, f'Branch "{name}" was deleted.')
        return redirect("dashboard")
    return render(request, "branches/delete_branch.html", {"branch": branch})


@staff_required
def export_branches_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Branches"
    ws.append(["branch_name", "contact_number", "address", "created_at", "updated_at"])
    for b in Branch.objects.all().iterator():
        ws.append(
            [
                b.branch_name,
                b.contact_number,
                b.address or "",
                b.created_at.isoformat() if b.created_at else "",
                b.updated_at.isoformat() if b.updated_at else "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(
        buf,
        as_attachment=True,
        filename="branches_export.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _normalize_header(h):
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def _import_rows(rows, request):
    """
    rows: iterable of dicts with keys branch_name, contact_number, optional address
    """
    created = 0
    skipped = 0
    errors = []
    for i, row in enumerate(rows, start=2):
        name = (row.get("branch_name") or "").strip()
        num = (row.get("contact_number") or "").strip()
        addr = (row.get("address") or "").strip()
        if not name and not num:
            skipped += 1
            continue
        if not name or not num:
            errors.append(f"Row {i}: branch_name and contact_number are required.")
            continue
        branch = Branch(branch_name=name, contact_number=num, address=addr)
        try:
            branch.full_clean()
            branch.save()
            created += 1
        except ValidationError as exc:
            error_text = "; ".join(
                [
                    str(message)
                    for messages in exc.message_dict.values()
                    for message in messages
                ]
            )
            errors.append(f"Row {i}: {error_text}")
    if created:
        messages.success(request, f"Imported {created} branch record(s).")
    if skipped:
        messages.info(request, f"Skipped {skipped} empty row(s).")
    for e in errors[:10]:
        messages.warning(request, e)
    if len(errors) > 10:
        messages.warning(request, f"...and {len(errors) - 10} more row errors.")


@staff_required
@require_http_methods(["POST"])
def import_branches(request):
    form = BranchImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Invalid file upload.")
        return redirect("dashboard")
    upload = form.cleaned_data["file"]
    name = (upload.name or "").lower()
    try:
        if name.endswith(".csv"):
            data = upload.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(data))
            if not reader.fieldnames:
                messages.error(request, "CSV has no header row.")
                return redirect("dashboard")
            fieldmap = {_normalize_header(h): h for h in reader.fieldnames if h}

            def col(raw, logical):
                orig = fieldmap.get(logical)
                return (raw.get(orig) or "").strip() if orig else ""

            rows = []
            for raw in reader:
                rows.append(
                    {
                        "branch_name": col(raw, "branch_name"),
                        "contact_number": col(raw, "contact_number"),
                        "address": col(raw, "address"),
                    }
                )
            with transaction.atomic():
                _import_rows(rows, request)
        elif name.endswith(".xlsx"):
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            headers = None
            body_rows = []
            for r in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [_normalize_header(c) for c in r]
                    continue
                if not any(r):
                    continue
                d = {}
                for j, key in enumerate(headers):
                    if key:
                        d[key] = r[j] if j < len(r) else ""
                body_rows.append(d)
            with transaction.atomic():
                _import_rows(body_rows, request)
        else:
            messages.error(request, "Please upload a .csv or .xlsx file.")
    except Exception as exc:  # noqa: BLE001 — surface import issues to admin
        messages.error(request, f"Import failed: {exc}")
    return redirect("dashboard")
